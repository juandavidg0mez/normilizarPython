import json
import base64
import logging
from datetime import datetime
from openpyxl import load_workbook
import os

# Configuración de logging
logging.basicConfig(level=logging.INFO)

def convertir_fechas(obj):
    """
    Serializador personalizado para objetos datetime.
    Convierte objetos datetime a formato ISO 8601.
    """
    if isinstance(obj, datetime):
        return obj.isoformat()
    raise TypeError(f"Tipo no serializable: {type(obj)}")

def lambda_handler(event, context):
    """
    Función principal de AWS Lambda para procesar archivos Excel codificados en Base64.
    Decodifica el archivo, extrae datos de las hojas y los estructura en un JSON.
    El JSON resultante se codifica en Base64 y se devuelve en el cuerpo de la respuesta.
    """
    try:
        body = event.get("body", None)
        if not body:
            logging.error("❌ Error: No se proporcionó un archivo Base64 en el evento.")
            return {
                "statusCode": 400,
                "body": json.dumps({"error": "No se proporcionó un archivo Base64 en el evento."})
            }

        # El cuerpo puede ser una cadena JSON, por lo que intentamos decodificarlo.
        if isinstance(body, str):
            try:
                body = json.loads(body)
            except json.JSONDecodeError:
                logging.error("❌ Error: Body no es un JSON válido.")
                return {
                    "statusCode": 400,
                    "body": json.dumps({"error": "Body vacío o inválido."})
                }
        
        # Verificar si el cuerpo está vacío o es inválido después de la decodificación
        if not body:
            logging.error("❌ Error: Body vacío o inválido después de decodificación.")
            return {
                "statusCode": 400,
                "body": json.dumps({"error": "Body vacío o inválido."})
            }

        archivo_base64 = body.get("file_base64", None)
        if not archivo_base64:
            logging.error("❌ Error: No se proporcionó un archivo Base64 válido.")
            return {
                "statusCode": 400,
                "body": json.dumps({"error": "No se proporcionó un archivo Base64 válido."})
            }

        # Decodificar el archivo Base64
        archivo_decodificado = base64.b64decode(archivo_base64)
        
        # Guardar el archivo decodificado temporalmente.
        # En Lambda, se usa /tmp/ para archivos temporales.
        ruta_temporal = os.path.join("/tmp", "archivo_temporal_para_procesar.xlsx")

        with open(ruta_temporal, "wb") as archivo:
            archivo.write(archivo_decodificado)

        # Cargar el libro de trabajo de Excel
        wb = load_workbook(filename=ruta_temporal, read_only=True, data_only=True)
        ArchivoPrincipal = {}

        # Procesar cada hoja del libro
        for hoja in wb.sheetnames:
            ws = wb[hoja]

            filas_crudas = []
            # Leer filas desde la 3 hasta la 214 y hasta la columna 50
            for row in ws.iter_rows(min_row=3, max_row=214, max_col=50, values_only=True):
                # Limpiar celdas vacías
                cleaned_row = [cell for cell in row if cell is not None and (isinstance(cell, str) and cell.strip() != "" or not isinstance(cell, str))]
                if cleaned_row:
                    filas_crudas.append(cleaned_row)

            datos_estructurados = {}
            seccion_actual = None
            seccion_id = 1
            current_table_headers = [] 
            is_in_table_section = False 
            
            for i, fila in enumerate(filas_crudas):
                # Detectar nuevas secciones principales (filas con un solo elemento)
                if len(fila) == 1: 
                    nombre = str(fila[0]).strip()
                    key = nombre.lower().replace(" ", "_")
                    
                    if key == "sin_seccion" and not seccion_actual: 
                        datos_estructurados.setdefault("sin_seccion", []).append(fila)
                        continue 
                        
                    if key in datos_estructurados:
                        key = f"{key}_{seccion_id}"
                        seccion_id += 1
                    datos_estructurados[key] = {}
                    seccion_actual = key
                    current_table_headers = [] 
                    is_in_table_section = False 
                    logging.info(f"✨ Nueva sección principal detectada: {key}")
                
                # Detectar cabeceras de tabla (filas con múltiples cadenas que preceden a datos)
                elif seccion_actual and len(fila) > 1 and all(isinstance(c, str) for c in fila) and i < len(filas_crudas) -1 :
                        next_row_is_likely_data = False
                        next_fila = filas_crudas[i+1] if i + 1 < len(filas_crudas) else []
                        
                        if next_fila and any(cell is not None and str(cell).strip() != "" for cell in next_fila):
                            if not (len(next_fila) == 1 and isinstance(next_fila[0], str) and next_fila[0].strip().lower().replace(" ", "_") not in ["sin_seccion"]):
                                 next_row_is_likely_data = True

                        if next_row_is_likely_data or (seccion_actual in ["error_de_relación_de_corriente_en_%_a_%_de_corriente_nominal", "fase_en_min_a_%_de_la_corriente_nominal", "datos_medidos"] and not current_table_headers):
                            current_table_headers = [str(cell).strip().lower().replace(" ", "_") for cell in fila]
                            is_in_table_section = True
                            if not isinstance(datos_estructurados.get(seccion_actual), list):
                                datos_estructurados[seccion_actual] = [] 
                            logging.info(f"📝 Cabeceras de tabla detectadas para {seccion_actual}: {current_table_headers}")
                            continue 

                # Procesar datos dentro de una sección
                if seccion_actual:
                    if is_in_table_section and current_table_headers:
                        row_data = {}
                        for idx, header in enumerate(current_table_headers):
                            if idx < len(fila):
                                row_data[header] = fila[idx]
                            else:
                                row_data[header] = None 
                        
                        if any(value is not None and str(value).strip() != "" for value in row_data.values()):
                            datos_estructurados[seccion_actual].append(row_data)
                            logging.info(f"📊 Fila de datos de tabla agregada a {seccion_actual}: {row_data}")
                        else:
                            logging.info(f"🚫 Fila de datos de tabla vacía, omitida: {fila}")

                    else:
                        # Procesar pares clave-valor dentro de una sección
                        subdata = {}
                        it = iter(fila)
                        try:
                            while True:
                                key_candidate = next(it)
                                value_candidate = next(it, None)
                                
                                if isinstance(key_candidate, (int, float)) or \
                                   (isinstance(key_candidate, str) and (len(str(key_candidate)) > 50 or \
                                   str(key_candidate).strip().lower() in ["ok", "si", "no", "desactivado", "protección", "ubicación", "colombia", "g3.2"] or \
                                   str(key_candidate).strip() == "" or \
                                   (value_candidate is None and not str(key_candidate).strip().lower().endswith(('_id', '_name', '_code'))))):
                                    
                                    logging.warning(f"⚠️ Posible clave no válida detectada: '{key_candidate}'. Añadiendo a 'valores_miscelaneos'.")
                                    if isinstance(datos_estructurados.get(seccion_actual), dict):
                                        datos_estructurados[seccion_actual].setdefault("valores_miscelaneos", []).extend([key_candidate, value_candidate])
                                    else: 
                                        datos_estructurados[seccion_actual].append({"valores_miscelaneos": [key_candidate, value_candidate]})

                                    continue 
                                    
                                key = str(key_candidate).strip().lower().replace(" ", "_")
                                if isinstance(datos_estructurados.get(seccion_actual), dict):
                                    subdata[key] = value_candidate
                                else: 
                                    datos_estructurados[seccion_actual].append({key: value_candidate})
                                    
                        except StopIteration:
                            pass
                        
                        if subdata: 
                            if isinstance(datos_estructurados.get(seccion_actual), dict):
                                datos_estructurados[seccion_actual].update(subdata)
                            
                # Si no hay sección actual, agregar a "sin_seccion"
                else: 
                    logging.info(f"❓ Fila sin sección asignada: {fila}")
                    subdata_sin_seccion = {}
                    is_key_value_pair = False
                    if len(fila) % 2 == 0: 
                        it = iter(fila)
                        temp_dict = {}
                        try:
                            while True:
                                key_candidate = next(it)
                                value_candidate = next(it, None)
                                if isinstance(key_candidate, str) and key_candidate.strip() != "":
                                    key = str(key_candidate).strip().lower().replace(" ", "_")
                                    temp_dict[key] = value_candidate
                                    is_key_value_pair = True
                                else:
                                    is_key_value_pair = False
                                    break
                        except StopIteration:
                            pass
                        if is_key_value_pair and temp_dict:
                            subdata_sin_seccion = temp_dict
                        else: 
                            subdata_sin_seccion = {"valores": fila}
                    else: 
                        subdata_sin_seccion = {"valores": fila}
                        
                    datos_estructurados.setdefault("sin_seccion", []).append(subdata_sin_seccion)

            ArchivoPrincipal[hoja] = datos_estructurados
        
        wb.close()
        # Eliminar el archivo temporal después de procesar
        if os.path.exists(ruta_temporal):
            os.remove(ruta_temporal)
            logging.info(f"🗑️ Archivo temporal '{ruta_temporal}' eliminado.")
        
        # Convertir el diccionario principal a una cadena JSON
        json_output_str = json.dumps(ArchivoPrincipal, default=convertir_fechas)
        
        # Codificar la cadena JSON en Base64
        encoded_json_output = base64.b64encode(json_output_str.encode('utf-8')).decode('utf-8')

        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': encoded_json_output,  # Retorna el JSON codificado en Base64
            'isBase64Encoded': True       # Indicar que el cuerpo está codificado en Base64
        }

    except Exception as e:
        logging.error(f"❌ Error: {str(e)}", exc_info=True) 
        return {
            'statusCode': 500,
            'body': json.dumps({"error": str(e)}),
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            }
        }